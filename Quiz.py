import openpyxl as op
from random import randint
from time import sleep
from Functions import mostrar_pergunta_respostas, verificar_resposta
from Functions import Ranking_Functions
from Functions import Menu_Quiz

# Main Program
quiz_book = op.load_workbook('Pasta1.xlsx')
sheet_quiz = quiz_book['Quiz']
sheet_cadastro = quiz_book['Cadastro']
sheet_admin = quiz_book['Admin']
sheet_rankings = quiz_book['Ranking']

# coloca os ranks em uma variavel lista
ranking = Ranking_Functions.carregando_ranking(sheet_rankings)

# Debug linhas perguntas
# linhas_totais = sheet_quiz.max_row
# print(linhas_totais)

jogador_admin = Menu_Quiz.Janela_principal(quiz_book, sheet_cadastro, sheet_admin)
# Apresentando o quiz
Menu_Quiz.Menu("QUIZ COPA DO MUNDO", jogador_admin[0])
# Determina o maximo de perguntas que existe na planilha
linhas = sheet_quiz.max_row - 1
Perguntas_Disponiveis = []
pontos = perguntas_respondidas = 0

# Loop onde o jogo so termina se o jogador quiser, ou não tiver mais nenhuma pergunta disponivel
while len(Perguntas_Disponiveis) != linhas:
    # Escolhe uma pergunta "aleatoriamente"
    num_perg = randint(1, linhas)
    # Se a pergunta não ja tiver sido escolhida, o programa procede com o jogo
    if num_perg not in Perguntas_Disponiveis:
        Perguntas_Disponiveis.append(num_perg)
        mostrar_pergunta_respostas(sheet_quiz, num_perg)
        while True:
            resposta = input('Qual a alternativa correta? ').upper().strip()
            # Verificando se a resposta do usuario é válida
            if resposta not in 'ABC' or resposta == '':
                print('\033[31m!! Digite uma alternativa válida !!\033[m')
                continue
            else:
                # Verifica a resposta do usuario é correda ou não
                verificador = verificar_resposta(sheet_quiz, num_perg, resposta)
                if verificador > 0:
                    if verificador == 2:
                        pontos += 2
                        perguntas_respondidas += 1
                    else:
                        pontos += 1
                        perguntas_respondidas += 1
                else:
                    perguntas_respondidas += 1
                break

    # Número minimo de perguntas, antes do usuario ter a opção de parar o jogo
    if len(Perguntas_Disponiveis) >= 5:
        # Após a primeira onda de perguntas, o programa irá perguntar se o jogador irá querer continuar.
        loop_continue = input('Deseja tentar mais uma pergunta?[S/N] ').strip().upper()
        while loop_continue not in 'SN' or loop_continue == '':
            print('\033[31m!! Digite uma alternativa válida !!\033[m')
            loop_continue = input('Deseja tentar mais uma pergunta?[S/N] ').strip().upper()

        # Caso o usuario não queira, o programa irá salvar os resultados da partida e então irá parar
        if loop_continue == 'N':
            print('Finalizando o QUIZ...')
            sleep(1)
            print('Registrando jogador...')
            # caucula os pontos ganhos na partida, e após isso adciona no perfil do usuario
            rank = Ranking_Functions.caucular_rank(pontos, perguntas_respondidas)
            jogador_admin[1] += rank
            print(f'\033[1;32mParabéns!! Você conseguiu \033[4m{rank}\033[m pontos!!\033[m')
            # atualiza o ranking da variavel, colocando o usuario na sua posiçao
            Ranking_Functions.alterando_ranking(ranking, jogador_admin)
            # salva esse novo ranking na planilha
            Ranking_Functions.alterando_ranking_planilha(sheet_rankings, ranking, quiz_book)
            Ranking_Functions.printando_ranking(ranking)
            sleep(1)
            print('QUIZ FINALIZADO.')
            break

        else:
            print('Criando uma nova pergunta...')
            sleep(1)

    else:
        print('Criando uma nova pergunta...')
        sleep(1)
